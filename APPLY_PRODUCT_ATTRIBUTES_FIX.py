from __future__ import annotations

from pathlib import Path
import shutil
import sys

ROOT = Path(__file__).resolve().parent
FILES = {
    "models": ROOT / "ocr_catalogue" / "models.py",
    "pipeline": ROOT / "ocr_catalogue" / "pipeline.py",
    "resolver": ROOT / "ocr_catalogue" / "offers" / "resolver.py",
    "index": ROOT / "static" / "index.html",
    "appjs": ROOT / "static" / "app.js",
    "exporter": ROOT / "ocr_catalogue" / "exporter.py",
    "tests_core": ROOT / "tests" / "test_core.py",
    "tests_offer": ROOT / "tests" / "test_offer_engine.py",
}


def backup(path: Path) -> None:
    backup_path = path.with_suffix(path.suffix + ".bak_product_attributes")
    if not backup_path.exists():
        shutil.copy2(path, backup_path)


def replace_once(text: str, old: str, new: str, label: str) -> str:
    count = text.count(old)
    if count != 1:
        raise RuntimeError(f"{label}: motif attendu 1 fois, trouvé {count} fois")
    return text.replace(old, new, 1)


def patch_models() -> None:
    path = FILES["models"]
    text = path.read_text(encoding="utf-8")
    old = '''    modele: str = ""\n    quantite: str = ""\n    prix_promo: str = ""'''
    new = '''    modele: str = ""\n    quantite: str = ""\n    caracteristiques: str = ""\n    prix_promo: str = ""'''
    text = replace_once(text, old, new, "Product.caracteristiques")
    path.write_text(text, encoding="utf-8")


def patch_pipeline() -> None:
    path = FILES["pipeline"]
    text = path.read_text(encoding="utf-8")
    old = '''        marque=offer.brand, modele=offer.model, quantite=offer.quantity,\n        prix_promo=offer.main_price, pourcentage=offer.percentage,'''
    new = '''        marque=offer.brand, modele=offer.model, quantite=offer.quantity,\n        caracteristiques=" • ".join(offer.technical_specs),\n        prix_promo=offer.main_price, pourcentage=offer.percentage,'''
    text = replace_once(text, old, new, "pipeline characteristics mapping")
    path.write_text(text, encoding="utf-8")


RESOLVER_HELPERS = r'''
# User-facing product attributes are separated from commercial pack format.
# This avoids treating 420 litres of refrigerator capacity or 7 kg of washing
# capacity as a retail "quantity" while keeping 2.5 litres of detergent in the
# Format / conditionnement field.
_APPLIANCE_PRODUCT = re.compile(
    r"\b(?:climatiseur|r[ée]frig[ée]rateur|cong[ée]lateur|lave[- ]?linge|machine\s+[àa]\s+laver|"
    r"lave[- ]?vaisselle|t[ée]l[ée]viseur|\btv\b|ventilateur|bouilloire|mixeur|batteur|blender|"
    r"hachoir|friteuse|cafeti[èe]re|fer\s+[àa]\s+repasser|s[èe]che[- ]?cheveux|p[èe]se\s+personne|"
    r"four|micro[- ]?ondes|cuisini[èe]re|r[ée]chaud|plaque|hotte|aspirateur|robot|grille[- ]?pain|"
    r"appareil\s+\d+\s+en\s+\d+)\b",
    re.I,
)
_STRONG_APPLIANCE_CUE = re.compile(
    r"\b(?:btu|watts?|\bw\b|tours?(?:/min)?|no\s+frost|inverter|tropicalis[ée]|garantie|"
    r"usb|hdmi|qled|oled|smart|r32|vitesses?|programmes?|couverts?|chaud\s*/\s*froid)\b",
    re.I,
)
_CHARACTERISTIC_CUE = re.compile(
    r"\b(?:btu|watts?|\bw\b|tours?(?:/min)?|hz|usb|hdmi|qled|oled|led|no\s+frost|inverter|"
    r"tropicalis[ée]|smart|r32|vitesses?|programmes?|couverts?|chaud\s*/\s*froid|garantie|"
    r"semelle|c[ée]ramique|inox|r[ée]cepteur|froid|capacit[ée])\b|"
    r"\b\d+(?:[,.]\d+)?\s*(?:btu|watts?|w|tours?|hz|usb|kg|litres?|l|cm|mm|pouces?|feux?)\b",
    re.I,
)
_CREDIT_ONLY = re.compile(r"\bachat\s+[àa]\s+cr[ée]dit\b|^\s*(?:\d+\s*mois\s*)+$", re.I)
_MODEL_AFTER_BRAND = re.compile(
    r"[”\"]\s*([A-Z0-9][A-Z0-9._/-]{2,})\b"
)


def _is_appliance_offer(product: str, objects: list[VisualObject]) -> bool:
    if _APPLIANCE_PRODUCT.search(product or ""):
        return True
    cue_lines = {
        obj.id for obj in objects
        if _STRONG_APPLIANCE_CUE.search(obj.text or "")
    }
    return len(cue_lines) >= 2


def _extract_model(objects: list[VisualObject]) -> str:
    explicit = [obj.text.strip() for obj in objects if obj.semantic_role == SemanticRole.MODEL and obj.text.strip()]
    if explicit:
        return explicit[0]
    # Monoprix frequently prints brand and model on the same physical line,
    # e.g. “MAXWELL” MX-CH12T-INV4-S. The line is semantically BRAND, so recover
    # the first alphanumeric reference immediately after the closing quote.
    for obj in objects:
        if obj.semantic_role != SemanticRole.BRAND:
            continue
        match = _MODEL_AFTER_BRAND.search(obj.text)
        if not match:
            continue
        token = match.group(1).strip("-./")
        if re.search(r"[A-Z]", token) and re.search(r"\d", token):
            return token
    return ""


def _dedupe_text(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        clean = re.sub(r"\s+", " ", value).strip(" -•")
        key = re.sub(r"\W+", "", clean).lower()
        if not clean or not key or key in seen:
            continue
        seen.add(key)
        result.append(clean)
    return result


def _format_and_characteristics(
    product: str,
    objects: list[VisualObject],
    facts: list[NumericFact],
) -> tuple[str, list[str]]:
    quantities = [obj.text for obj in objects if obj.semantic_role == SemanticRole.QUANTITY]
    appliance = _is_appliance_offer(product, objects)

    characteristics: list[str] = []
    for obj in objects:
        if obj.semantic_role != SemanticRole.TECHNICAL_SPEC:
            continue
        text = obj.text.strip()
        if not text or _CREDIT_ONLY.search(text):
            continue
        # Do not expose generic variant copy as a technical characteristic.
        if appliance or _CHARACTERISTIC_CUE.search(text):
            characteristics.append(text)

    # Numeric facts are kept as evidence even when line classification was
    # imperfect. This recovers BTU, power, dimensions, duration, etc.
    for fact in facts:
        if fact.role not in {
            NumericRole.POWER,
            NumericRole.CAPACITY,
            NumericRole.DIMENSION,
            NumericRole.DURATION,
            NumericRole.TECHNICAL_SPEC,
        }:
            continue
        text = fact.text.strip()
        if not text or _CREDIT_ONLY.search(text):
            continue
        if fact.role == NumericRole.DURATION and not re.search(r"garantie", text, re.I):
            # Bare 18/36 months on these catalogues is usually credit duration.
            continue
        characteristics.append(text)

    if appliance:
        # Capacity-like lines are technical on appliances: 12000 BTU,
        # refrigerator 420 L, washing-machine 7 kg, freezer 84 L, etc.
        characteristics.extend(quantities)
        retail_format = ""
    else:
        # Food, cleaning, hygiene and other FMCG keep their commercial pack.
        retail_format = quantities[-1] if quantities else ""

    return retail_format, _dedupe_text(characteristics)

'''


def patch_resolver() -> None:
    path = FILES["resolver"]
    text = path.read_text(encoding="utf-8")
    marker = '''def _assemble(page: PageScene, candidates: list[OfferCandidate], graph: SpatialGraph, style, region_solutions: dict[str, RegionSolution]) -> list[Offer]:\n'''
    text = replace_once(text, marker, RESOLVER_HELPERS + marker, "insert product attribute helpers")

    old = '''        quantities = [obj.text for obj in objects if obj.semantic_role == SemanticRole.QUANTITY]\n        models = [obj.text for obj in objects if obj.semantic_role == SemanticRole.MODEL]\n        technical = [obj.text for obj in objects if obj.semantic_role == SemanticRole.TECHNICAL_SPEC]\n        cashback = next((fact for fact in facts if fact.role == NumericRole.CASHBACK), None)'''
    new = '''        quantity, technical = _format_and_characteristics(product, objects, facts)\n        model = _extract_model(objects)\n        cashback = next((fact for fact in facts if fact.role == NumericRole.CASHBACK), None)'''
    text = replace_once(text, old, new, "resolver split format and characteristics")

    old = '''            brand=next((brand for brand in brands if brand), ""), model=models[0] if models else "",\n            quantity=quantities[-1] if quantities else "", main_price=(main.value + " DT") if main else "",'''
    new = '''            brand=next((brand for brand in brands if brand), ""), model=model,\n            quantity=quantity, main_price=(main.value + " DT") if main else "",'''
    text = replace_once(text, old, new, "resolver output fields")
    path.write_text(text, encoding="utf-8")


def patch_index() -> None:
    path = FILES["index"]
    text = path.read_text(encoding="utf-8")
    old = '''<th>Photo</th><th>Produit</th><th>Marque</th><th>Quantité</th><th>Prix promo</th><th>Pourcentage</th><th>Promotion</th>'''
    new = '''<th>Photo</th><th>Produit</th><th>Marque</th><th>Modèle</th><th>Format / conditionnement</th><th>Caractéristiques</th><th>Prix promo</th><th>Pourcentage</th><th>Promotion</th>'''
    text = replace_once(text, old, new, "rename/add UI columns")
    path.write_text(text, encoding="utf-8")


def patch_appjs() -> None:
    path = FILES["appjs"]
    text = path.read_text(encoding="utf-8")
    old = "${['produit','marque','quantite','prix_promo','pourcentage','promotion'].map(key=>editableCell(product,key)).join('')}"
    new = "${['produit','marque','modele','quantite','caracteristiques','prix_promo','pourcentage','promotion'].map(key=>editableCell(product,key)).join('')}"
    text = replace_once(text, old, new, "UI editable columns")
    path.write_text(text, encoding="utf-8")


def patch_exporter() -> None:
    path = FILES["exporter"]
    text = path.read_text(encoding="utf-8")
    text = replace_once(
        text,
        "from openpyxl.styles import Font, PatternFill\n",
        "from openpyxl.styles import Font, PatternFill\nfrom openpyxl.utils import get_column_letter\n",
        "exporter get_column_letter import",
    )
    old = '''COLUMNS = [\n    ("photo", "Photo"), ("produit", "Produit"), ("marque", "Marque"),\n    ("quantite", "Quantité"), ("prix_promo", "Prix promo"),\n    ("pourcentage", "Pourcentage"),\n    ("promotion", "Promotion"), ("page", "Page"),\n    ("confiance", "Confiance"), ("statut", "Statut"),\n]'''
    new = '''COLUMNS = [\n    ("photo", "Photo"), ("produit", "Produit"), ("marque", "Marque"),\n    ("modele", "Modèle"), ("quantite", "Format / conditionnement"),\n    ("caracteristiques", "Caractéristiques"), ("prix_promo", "Prix promo"),\n    ("pourcentage", "Pourcentage"), ("promotion", "Promotion"),\n    ("page", "Page"), ("confiance", "Confiance"), ("statut", "Statut"),\n]'''
    text = replace_once(text, old, new, "export columns")

    old = '''    widths = {"A": 14, "B": 34, "C": 20, "D": 18, "E": 16, "F": 16, "G": 12, "H": 28}\n    for key, width in widths.items():\n        ws.column_dimensions[key].width = width'''
    new = '''    widths = {\n        "Photo": 14, "Produit": 30, "Marque": 18, "Modèle": 22,\n        "Format / conditionnement": 24, "Caractéristiques": 48,\n        "Prix promo": 16, "Pourcentage": 16, "Promotion": 28,\n        "Page": 10, "Confiance": 12, "Statut": 14,\n    }\n    for index, (_, label) in enumerate(columns, start=1):\n        ws.column_dimensions[get_column_letter(index)].width = widths.get(label, 16)'''
    text = replace_once(text, old, new, "dynamic export widths")
    path.write_text(text, encoding="utf-8")


def patch_tests() -> None:
    core = FILES["tests_core"]
    text = core.read_text(encoding="utf-8")
    old = '''        self.assertIn("Prix promo", headers)\n        self.assertIn("Pourcentage", headers)\n        self.assertNotIn("Ancien prix", headers)'''
    new = '''        self.assertIn("Prix promo", headers)\n        self.assertIn("Pourcentage", headers)\n        self.assertIn("Format / conditionnement", headers)\n        self.assertIn("Caractéristiques", headers)\n        self.assertIn("Modèle", headers)\n        self.assertNotIn("Quantité", headers)\n        self.assertNotIn("Ancien prix", headers)'''
    text = replace_once(text, old, new, "export header regression test")
    core.write_text(text, encoding="utf-8")

    offer = FILES["tests_offer"]
    text = offer.read_text(encoding="utf-8")
    old_import = '''from ocr_catalogue.offers.resolver import _merge_product_with_priced_brand, _offer_bbox, _offer_candidates, _partition_container, _pick_product, _reassign_secondary_facts\n'''
    new_import = '''from ocr_catalogue.offers.resolver import _extract_model, _format_and_characteristics, _merge_product_with_priced_brand, _offer_bbox, _offer_candidates, _partition_container, _pick_product, _reassign_secondary_facts\n'''
    text = replace_once(text, old_import, new_import, "offer test helper imports")

    marker = '''    def test_classifier_accepts_only_explicit_free_mechanism(self):\n'''
    tests = '''    def test_appliance_capacity_moves_to_characteristics_not_format(self):\n        product = VisualObject("product", 1, "line", BBox(0, 0, 90, 12), text="Réfrigérateur 2 portes", semantic_role=SemanticRole.PRODUCT_TEXT)\n        capacity = VisualObject("capacity", 1, "line", BBox(0, 20, 70, 32), text="420 LITRES", semantic_role=SemanticRole.QUANTITY)\n        frost = VisualObject("frost", 1, "line", BBox(0, 40, 80, 52), text="NO FROST", semantic_role=SemanticRole.TECHNICAL_SPEC)\n        retail_format, characteristics = _format_and_characteristics("Réfrigérateur 2 portes", [product, capacity, frost], [])\n        self.assertEqual(retail_format, "")\n        self.assertIn("420 LITRES", characteristics)\n        self.assertIn("NO FROST", characteristics)\n\n    def test_fmcg_volume_stays_in_format(self):\n        product = VisualObject("product", 1, "line", BBox(0, 0, 90, 12), text="Lessive liquide machine", semantic_role=SemanticRole.PRODUCT_TEXT)\n        volume = VisualObject("volume", 1, "line", BBox(0, 20, 70, 32), text="2,35 litres", semantic_role=SemanticRole.QUANTITY)\n        retail_format, characteristics = _format_and_characteristics("Lessive liquide machine", [product, volume], [])\n        self.assertEqual(retail_format, "2,35 litres")\n        self.assertEqual(characteristics, [])\n\n    def test_btu_is_exposed_as_appliance_characteristic(self):\n        product = VisualObject("product", 1, "line", BBox(0, 0, 90, 12), text="Climatiseur", semantic_role=SemanticRole.PRODUCT_TEXT)\n        btu = VisualObject("btu", 1, "line", BBox(0, 20, 80, 32), text="12000 BTU", semantic_role=SemanticRole.TECHNICAL_SPEC)\n        _, characteristics = _format_and_characteristics("Climatiseur", [product, btu], [])\n        self.assertIn("12000 BTU", characteristics)\n\n    def test_model_is_recovered_from_brand_line(self):\n        brand = VisualObject("brand", 1, "line", BBox(0, 0, 180, 12), text='“MAXWELL” MX-CH12T-INV4-S', semantic_role=SemanticRole.BRAND)\n        self.assertEqual(_extract_model([brand]), "MX-CH12T-INV4-S")\n\n'''
    text = replace_once(text, marker, tests + marker, "offer attribute regression tests")
    offer.write_text(text, encoding="utf-8")


def main() -> None:
    missing = [str(path) for path in FILES.values() if not path.exists()]
    if missing:
        raise RuntimeError("Fichiers introuvables:\n" + "\n".join(missing))

    for path in FILES.values():
        backup(path)

    patch_models()
    patch_pipeline()
    patch_resolver()
    patch_index()
    patch_appjs()
    patch_exporter()
    patch_tests()

    print("PRODUCT ATTRIBUTES FIX APPLIQUE")
    print("Colonnes UI: Modèle | Format / conditionnement | Caractéristiques")
    print("Les capacités d'électroménager (BTU, litres, kg, etc.) vont dans Caractéristiques.")
    print("Les formats FMCG (2,5 L de lessive, 400 ml shampooing, lots...) restent dans Format / conditionnement.")
    print(r"Etape suivante: .\test.ps1")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERREUR: {exc}", file=sys.stderr)
        raise
